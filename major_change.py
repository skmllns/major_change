from collections import Counter
import time

from openpyxl import load_workbook

# from openpyxl.cell import column_index_from_string

read_file = 'Major-School-College-Change 8.xlsx'
write_file = 'table.txt'

start_time = time.time()
wb = load_workbook(filename=read_file, read_only=True)
sheet = wb[wb.get_sheet_names()[0]]

row_count = sheet.max_row

# 6 semester major sequence code: BJ, 62

# convert to google charts format
with open(write_file, 'wb') as wf:
   two_letters = {}
   for row_num in range(2, row_count + 1):
      code = str(sheet.cell(row=row_num, column=64).value[0]) + str(sheet.cell(row=row_num, column=64).value[1])
      two_letters[code] = two_letters.get(code, 0) + 1
   for key in two_letters:
      key_len = len(key)
      str_to_write = '['
      for idx in range(key_len):
         if idx + 1 < key_len:
            str_to_write += '\'' + key[idx] + str(idx + 1) + '\', \'' + key[idx + 1] + str(idx + 2) + '\', '
         str_to_write += str(two_letters[key]) + '],\n'
         print str_to_write
                
      # wf.writelines(['[\'' + key[0] + '\', \'' + elem + '\', ' + str(all_codes[key]) + '],\n'])
